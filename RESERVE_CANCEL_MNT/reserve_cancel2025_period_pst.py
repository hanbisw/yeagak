from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.QtCore import *
from PyQt5.QtGui import QBrush, QRegExpValidator, QColor, QPalette, QFont
from PyQt5.QtWidgets import *

import pandas as pd
import subprocess
import webbrowser

from config.mariadb_connection import mariadb_conn
import mariadb
from datetime import timedelta,datetime
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,PatternFill,Border,Side,colors
from pathlib import Path
import os
import threading
import time

# class StockIndexWindow(QMainWindow, form_class):
class Reserve_cancel2025_period_pst_Window(QMainWindow):
    def __init__(self,pf_os,comp_code,user_id):
        super().__init__()
        self.move(10, 50)
        self.setFixedSize(1910, 903)
        if pf_os == 'Windows':
            option_ui = 'UiDir/reserve_cancel2025_period_pst.ui'
        else:
            option_ui = 'UiDir/reserve_cancel2025_period_pst.ui'


        self.comp_code = comp_code
        self.user_id = user_id
        # option_ui = 'UiDir/StockCompInfo.ui'
        uic.loadUi(option_ui, self)

        regExp = QRegExp("[0-9]*") #edit에 숫자만 입력 처리하기 위해
        palette = QPalette()
        palette.setColor(QPalette.Highlight, QColor(144, 153, 234))  # default ==> Qt.darkBlue
        # palette.setColor(QPalette.HighlightedText, Qt.red)  # default ==> Qt.white
        self.tw_reserve.setPalette(palette)

        self.ed_cancel_sum.setValidator(QRegExpValidator(regExp, self))#숫자 만 입력 하려면.....
        self.str_ = '-'  # 203-1호실일때 -호일이 있는가를 찾으려면.....
        self.sch_room_name = ''  # 첫 화면시 호실 정보가 없는 경우에.....
        self.init_rtn()#오늘 일자를 먼저 보여주려면......
        self.pb_search.clicked.connect(self.select_rtn)  # 자료 조회
        self.pb_clear.clicked.connect(self.reserve_clear)  # 저장 화면을 클리어
        self.pb_excel.clicked.connect(self.excel_rtn)  #엑셀 파일로 저장 어
        self.tw_reserve.clicked.connect(self.mouse_click_rtn) # 클릭시 해당일자 호실의 정보를 상당 입력란에 보여주려면.....
        self.tw_reserve.keyPressEvent = self.keyPressEvent #UP, DOWN 키를 누를 때 해당일자 호실의 정보를 상당 입력란에 보여주려면.....
        self.first_sch()  # 첫 화면시 현재일자를 검색해온다.

    #오늘 일자를 먼저 보여주려면....
    def init_rtn(self):

        # 현재 일자를 검색
        mra = mariadb_conn().conn
        csr = mra.cursor()
        query_sel = ("SELECT date_format(date_add(CURDATE(), INTERVAL -1  day),'%Y-%m-%d'),"
                     "date_format(CURDATE(),'%Y-%m-%d'),CURDATE() AS ILJA,DATE_ADD(CURDATE(), INTERVAL 1  day) from dual ")
        # print(query_sel)
        csr.execute(query_sel)
        rows = csr.fetchall()
        csr.close()
        mra.close()
        self.yesterday_chk = rows[0][0]
        start_ilja = rows[0][0]
        end_ilja = rows[0][1]
        self.now_date = str(rows[0][2])  # 오늘 일자를 기준으로 자료의 입력을 결정하기 위해

        self.de_start_ilja_sch.setDate(datetime.strptime(start_ilja, '%Y-%m-%d'))  # 기간 조회 시작일자
        self.de_end_ilja_sch.setDate(datetime.strptime(end_ilja, '%Y-%m-%d'))  # 기간 조회 종료일자

        # 예약 사이트 정보 조회
        mra = mariadb_conn().conn
        csr = mra.cursor(dictionary=True) #딕셔너리 형태로 표현하기 위해....
        # query_sel = "SELECT APP_COMP_CODE, APP_COMP_NAME FROM app_company_info_tbl ORDER BY APP_COMP_CODE "
        query_sel = ("SELECT B.APPCOMP_NAME "
                     " FROM appcomp_use_info_tbl a JOIN appcomp_info_tbl b USING(appcomp_code)"
                     " WHERE comp_code = '"+self.comp_code+"' ORDER BY A.APPCOMP_CODE ")
        # print(query_sel)
        csr.execute(query_sel)
        rows = csr.fetchall()
        csr.close()
        mra.close()
        self.cb_site_sch.insertItem(0,'전체')
        #예약 사이트 정보
        for idx, col in enumerate(rows):
            self.cb_site_sch.insertItem(idx+1,col["APPCOMP_NAME"])

    #조회시 일자 값이 정확한지 체크
    def select_rtn(self):
        if self.de_start_ilja_sch.date() > self.de_end_ilja_sch.date():
            QMessageBox.about(self, "정보", "일자를 정확히 입력하세요....")
            self.tw_reserve.setRowCount(0)
            self.de_start_ilja_sch.setFocus()
        else:
            self.first_sch()

    # 클릭시 층수를 ed_floor_st edit에 보여준다....
    def reserve_clear(self):
        self.ed_name.setText("")  # 예약자
        self.ed_reserve_number.setText("")  # 예약번호
        self.ed_room_name.setText("")  # 호실
        self.ed_room_type.setText("")  # 호실 타입
        self.ed_sum.setText("")  # 결재 금액
        self.ed_select_ilja.setText("")  # 선택일자
        self.ed_worker.setText("")  # 당일 근무자
        self.ed_bigo.setText("")  # 예약자
        self.ed_addsum.setText("")  # 추가 금액
        self.ed_addbigo.setText("")  # 추가결재 비고
        self.ed_reserve_ilja.setText("")#예약일자
        self.ed_start_ilja.setText("")#입실일자
        self.ed_end_ilja.setText("")#퇴실일자

        self.ed_reserve_site.setText("") #예약 사이트 다시 가져오기 위해 초기화
        self.ed_payment.setText("") #결재수단 정보 다시 가져오기 위해 초기화
        self.ed_addpayment.setText("") #추가 결재수단 정보 다시 가져오기 위해 초기화
        self.ed_cancel_start_ilja.setText("")#예약 현재일을 가져와 취소 시작일자로 한다. 10박중 5박만 취소하는 경우....
        self.ed_cancel_end_ilja.setText("")#퇴실일자는 그래로 가져와 마지막을 표시한다.
        self.ed_mnt_id.setText(self.user_id)#로그인 id를 보여주고 취소 자료 처리를 누가 했는지를 알기 위해......
        self.ed_cancel_sum.setText("")#환불 금액을 초기화
        self.ed_cancel_bigo.setText("")#예약 취소 사유를 초기화
        self.ed_cancel_payment.setText("")#환불 결재수단 정보 초기화

        self.ed_mnt_id.setText("")# 취소자 ID 초기화
        self.ed_cancel_start_ilja.setText("")# 취소사항 시작일자 초기화
        self.ed_cancel_end_ilja.setText("")# 취소사항 종료일자 초기화
        self.ed_cancel_payment.setText("")# 취소사항 환불수단 초기화
        self.ed_cancel_sum.setText("") #환불금액 초기화
        self.ed_cancel_bigo.setText("") #취사 참고사항 초기화

    #KEY_UP, KEY_DOWN을 눌렀다면 .....
    def keyPressEvent (self, eventQKeyEvent):
        key = eventQKeyEvent.key()
        if key == Qt.Key_Up: #up 키를 누를 때마다  해당일자 호실의 정보를 상당 입력란에 보여주려면.....
            row = self.tw_reserve.currentRow() #현재 ROW 값을 가져오려면..
            if row == 0: # 처음 일때 값을 1로 하기 위해
                row += 1
            self.tw_reserve.setCurrentCell(row-1,0)  #한 ROW 위로 이동
            sch_click_ilja = self.tw_reserve.item(self.tw_reserve.currentRow(), 0).text()  #해당일자를 읽는다.
            sch_reserve_name = self.tw_reserve.item(self.tw_reserve.currentRow(), 4).text()  # 예약 번호을 읽는다.
            sch_reserve_number = self.tw_reserve.item(self.tw_reserve.currentRow(), 7).text()  # 예약 번호을 읽는다.
            # self.reserve_clear()
            self.search_rtn(sch_click_ilja,sch_reserve_number,sch_reserve_name)#해당 일자의 정보를 화면에 보여준다.

        elif key == Qt.Key_Down:  # down 키를 누를 때마다  해당일자 호실의 정보를 상당 입력란에 보여주려면.....
            row = self.tw_reserve.currentRow() #현재 ROW 값을 가져오려면..
            if row >= (self.tw_reserve.rowCount() -1): #마지막 ROW일때 계속 마지막 해당일자 호실의 정보를 상당 입력란에 보여주려면.....
               row -= 1
            self.tw_reserve.setCurrentCell(row + 1, 0) #한 ROW 밑으로 이동
            sch_click_ilja = self.tw_reserve.item(self.tw_reserve.currentRow(), 0).text()  #해당일자를 읽는다.
            sch_reserve_name = self.tw_reserve.item(self.tw_reserve.currentRow(), 4).text()  # 예약 번호을 읽는다.
            sch_reserve_number = self.tw_reserve.item(self.tw_reserve.currentRow(), 7).text()  # 예약 번호을 읽는다.
            # self.reserve_clear()
            self.search_rtn(sch_click_ilja,sch_reserve_number,sch_reserve_name)#해당 일자의 정보를 화면에 보여준다.


    # 클릭시 해당일자 호실의 정보를 상당 입력란에 보여주려면.....
    def mouse_click_rtn(self):
        row = self.tw_reserve.currentRow() #현재 ROW 값을 가져오려면..
        self.tw_reserve.setCurrentCell(row, 0)  # ROW 이동
        sch_click_ilja = self.tw_reserve.item(self.tw_reserve.currentRow(), 0).text()  #해당일자를 읽는다.
        sch_reserve_name = self.tw_reserve.item(self.tw_reserve.currentRow(), 4).text()  # 예약 번호을 읽는다.
        sch_reserve_number = self.tw_reserve.item(self.tw_reserve.currentRow(), 7).text()  # 예약 번호을 읽는다.
        # self.reserve_clear()
        self.search_rtn(sch_click_ilja,sch_reserve_number,sch_reserve_name)#해당 일자의 정보를 화면에 보여준다.

    #해당일자 호실에 대한 정보를 상단 입력란에 뿌려주려면.....
    def search_rtn(self,sch_click_ilja,sch_reserve_number,sch_reserve_name):

        #일자별 호실 정보 가져오기 60일 치
        mra = mariadb_conn().conn
        csr = mra.cursor(dictionary=True) #딕셔너리 형태로 표현하기 위해....
        query_sel =("SELECT DATE_FORMAT(a.ROOM_ILJA,'%Y-%m-%d') as ILJA, A.room_name AS RN,"
                    "(SELECT room_type from yeogak_room_name_info_tbl where comp_code = a.comp_code and room_name = a.room_name) AS ROOM_TYPE,"
                    " A.RESERVE_NAME,decode_oracle(A.RESERVE_SITE,'','',FN_APPCOMP_NAME(A.RESERVE_SITE)) AS SITE_NAME, "
                    "A.RESERVE_ILJA,A.RESERVE_NUMBER,A.RESERVE_START_ILJA,A.RESERVE_END_ILJA,"
                    "decode_oracle(a.payment,'','',null,'',FN_CODE_NAME('PAY',a.payment)) AS PAYMENT_NAME,"
                    "FORMAT(a.PAY_SUM,0) AS PAY_SUM,"
                    "FORMAT(a.EXPEC_SUM,0) AS EXPEC_SUM, "
                    "DECODE_ORACLE(A.DORMITORY_SEX,'g','여자','m','남자','') AS DORMITORY_SEX,DECODE_ORACLE(A.DORMITORY_CLOSE,'c','Yes','') AS DORMITORY_CLOSE, "
                    "A.WORKER,A.BIGO, " 
                    "(SELECT decode_oracle(add_payment,'', '',FN_CODE_NAME('PAY',add_payment)) from yeogak_addpayment2025_tbl WHERE comp_code = b.comp_code "
                    "AND key_ilja = a.key_ilja AND room_ilja =a.room_ilja) AS ADD_PAYMENT, "
                    "(SELECT decode_oracle(add_sum, NULL, '',FORMAT(add_sum,0)) "
                    "from yeogak_addpayment2025_tbl WHERE comp_code = a.comp_code AND key_ilja = a.key_ilja AND room_ilja =a.room_ilja) AS ADD_SUM, "
                    "(SELECT decode_oracle(add_bigo, NULL,'',add_bigo) from yeogak_addpayment2025_tbl WHERE comp_code = a.comp_code "
                    "AND key_ilja = a.key_ilja AND room_ilja =a.room_ilja) AS ADD_BIGO, "
                    "datediff(date_format(a.reserve_end_ilja, '%Y-%m-%d'),date_format(a.reserve_start_ilja, '%Y-%m-%d')) AS DATE_CHA, "                      
                    "B.CANCEL_START_ILJA, "
                    "B.CANCEL_END_ILJA,DECODE_ORACLE(B.CANCEL_PAYMENT,'','',FN_CODE_NAME('CPY',B.CANCEL_PAYMENT))  AS CANCEL_PAYMENT_NAME, "
                    "FORMAT(B.CANCEL_SUM,0) AS CANCEL_SUM, "
                    "B.CANCEL_BIGO,B.MNT_ID,A.SPECIAL_NOTE "
                    "FROM  yeogak_reserve2025_tbl a, yeogak_reserve2025_cancel_tbl b "   
                    "where a.comp_code = %s and a.room_ilja = %s  and a.reserve_number = %s  and a.reserve_name = %s  AND a.cancel_chk = '1' " 							
                    "and a.comp_code = b.comp_code and a.key_ilja = b.key_ilja AND a.reserve_number = b.reserve_number "
                    "ORDER BY a.room_ilja,a.reserve_name ")
        # print(query_sel)
        t = (self.comp_code,sch_click_ilja,sch_reserve_number,sch_reserve_name)
        # print(t)
        csr.execute(query_sel,t)
        rows = csr.fetchall()
        csr.close()
        mra.close()

        self.ed_room_name.setText(rows[0]['RN'])  # 호실
        self.ed_room_type.setText(rows[0]['ROOM_TYPE'])  # 호실 타입
        self.ed_name.setText(rows[0]['RESERVE_NAME'])  # 예약자
        self.ed_reserve_site.setText(rows[0]['SITE_NAME']) #예약 사이트 다시 가져오기 위해 초기화
        self.ed_reserve_ilja.setText(rows[0]['RESERVE_ILJA'])  # 예약일자
        self.ed_reserve_number.setText(rows[0]['RESERVE_NUMBER'])  # 예약번호
        self.ed_select_ilja.setText(rows[0]['RESERVE_ILJA'])# 선택일자
        self.ed_expec_sum.setText(rows[0]['EXPEC_SUM'])  # 예상 금액
        self.ed_start_ilja.setText(rows[0]['RESERVE_START_ILJA'])  # 입실일자
        self.ed_end_ilja.setText(rows[0]['RESERVE_END_ILJA'])  # 퇴실일자
        self.ed_payment.setText(rows[0]['PAYMENT_NAME'])#결재수단 정보 다시 가져오기 위해 초기화
        self.ed_sum.setText(rows[0]['PAY_SUM'])  # 결재 금액
        self.ed_special_note.setText(rows[0]['SPECIAL_NOTE'])  # 특이사항
        self.ed_worker.setText(rows[0]['WORKER'])  # 당일 근무자
        self.ed_bigo.setText(rows[0]['BIGO'])  # 요구사항

        self.ed_addpayment.setText(rows[0]['ADD_PAYMENT']) #추가 결재수단 정보 다시 가져오기 위해 초기화
        self.ed_addsum.setText(rows[0]['ADD_SUM'])  # 추가 금액
        self.ed_addbigo.setText(rows[0]['ADD_BIGO'])  # 추가결재 비고
        sch_room_name= rows[0]['RN']
        #203-1 '-' 있는지를 체크하여 활성화 및 비활성화하려면...
        if sch_room_name.find(self.str_) != -1:
            self.cb_close.setVisible(True)# 마감 체크 보여주려면
            self.lb_sex.setVisible(True)#남/여 호실 구분 보여주려면
            self.ed_sex.setVisible(True)# 남/여 호실 보여주려면
        else:
            self.cb_close.setVisible(False)# 마감 체크 안 보여주려면
            self.lb_sex.setVisible(False)#남/여 호실 구분 안 보여주려면
            self.ed_sex.setVisible(False)# 남/여 호실 안 보여주려면
        self.ed_cancel_start_ilja.setText(rows[0]['CANCEL_START_ILJA'])# 취소사항 시작일자
        self.ed_cancel_end_ilja.setText(rows[0]['CANCEL_END_ILJA'])# 취소사항 종료일자
        self.ed_cancel_payment.setText(rows[0]['CANCEL_PAYMENT_NAME'])# 취소사항 환불수단
        self.ed_cancel_sum.setText(rows[0]['CANCEL_SUM']) #환불금액
        self.ed_cancel_bigo.setText(rows[0]['CANCEL_BIGO'])  #취사 참고사항
        self.ed_mnt_id.setText(rows[0]['MNT_ID'])# 취소자 ID

    #자료를 가져와 화면에 뿌려준다.
    def first_sch(self):
        self.reserve_clear()
        #사이트 조회
        site_name = self.cb_site_sch.currentText()#사이트

        if self.cb_site_sch.currentText() == '전체':
            site_code = '%'
        else:
            #코드 값 가져오기
            mra = mariadb_conn().conn
            csr = mra.cursor()
            query_sel = "SELECT FN_APPCOMP_CODE('"+site_name+"') "
            # print(query_sel)
            csr.execute(query_sel)
            rows = csr.fetchall()
            csr.close()
            mra.close()
            site_code = rows[0][0] #site code값을 가져온다

        start_ilja_sch = self.de_start_ilja_sch.date().toString('yyyy-MM-dd')#입실일자
        end_ilja_sch = self.de_end_ilja_sch.date().toString('yyyy-MM-dd')#퇴실일자


        #일자별 호실 정보 가져오기 60일 치
        mra = mariadb_conn().conn
        csr = mra.cursor(dictionary=True) #딕셔너리 형태로 표현하기 위해....
        query_sel =("SELECT DATE_FORMAT(a.ROOM_ILJA,'%Y-%m-%d') as ILJA,(select WK from yeogak_reserve_day_tbl where ilja = a.room_ilja) AS WK,A.room_name AS RN, "
                    "(SELECT room_type from yeogak_room_name_info_tbl where comp_code = a.comp_code and room_name = a.room_name) "
                    "AS ROOM_TYPE, A.RESERVE_NAME,decode_oracle(A.RESERVE_SITE,'','',FN_APPCOMP_NAME(A.RESERVE_SITE)) AS SITE_NAME, "
                    "A.RESERVE_ILJA,A.RESERVE_NUMBER,A.RESERVE_START_ILJA,A.RESERVE_END_ILJA, decode_oracle(a.payment,'','',null,'',FN_CODE_NAME('PAY',a.payment)) AS PAYMENT_NAME, FORMAT(a.PAY_SUM,0) AS PAY_SUM,FORMAT(a.EXPEC_SUM,0) AS EXPEC_SUM, "
                    "DECODE_ORACLE(A.DORMITORY_SEX,'g','여자','m','남자','') AS DORMITORY_SEX,DECODE_ORACLE(A.DORMITORY_CLOSE,'c','Yes','') AS DORMITORY_CLOSE, "
                    "A.WORKER,A.BIGO, " 
                    "(SELECT decode_oracle(add_payment,'', '',FN_CODE_NAME('PAY',add_payment)) from yeogak_addpayment2025_tbl WHERE comp_code = b.comp_code "
                    "AND key_ilja = a.key_ilja AND room_ilja =a.room_ilja ) AS ADD_PAYMENT, "
                    "(SELECT decode_oracle(add_sum, NULL, '',FORMAT(add_sum,0)) "
                    "from yeogak_addpayment2025_tbl WHERE comp_code = a.comp_code AND key_ilja = a.key_ilja AND room_ilja =a.room_ilja ) AS ADD_SUM, "
                    "(SELECT decode_oracle(add_bigo, NULL,'',add_bigo) from yeogak_addpayment2025_tbl WHERE comp_code = a.comp_code "
                    "AND key_ilja = a.key_ilja AND room_ilja =a.room_ilja) AS ADD_BIGO, "
                    "datediff(date_format(a.reserve_end_ilja, '%Y-%m-%d'),date_format(a.reserve_start_ilja, '%Y-%m-%d')) AS DATE_CHA, "                      
                    "B.CANCEL_START_ILJA, "
                    "B.CANCEL_END_ILJA,DECODE_ORACLE(B.CANCEL_PAYMENT,'','',FN_CODE_NAME('CPY',B.CANCEL_PAYMENT))  AS CANCEL_PAYMENT_NAME, "
                    "FORMAT(B.CANCEL_SUM,0) AS CANCEL_SUM, "
                    "B.CANCEL_BIGO,B.MNT_ID,A.SPECIAL_NOTE "
                    "FROM  yeogak_reserve2025_tbl a, yeogak_reserve2025_cancel_tbl b "   
                    "where a.comp_code = %s and a.room_ilja BETWEEN %s AND %s  and a.reserve_site LIKE %s AND a.cancel_chk = '1' " 							
                    "and a.comp_code = b.comp_code and a.key_ilja = b.key_ilja AND a.reserve_number = b.reserve_number "
                    "ORDER BY a.room_ilja,a.reserve_name ")
        # print(query_sel)
        t = (self.comp_code,start_ilja_sch,end_ilja_sch,site_code)
        csr.execute(query_sel,t)
        rows = csr.fetchall()
        csr.close()
        mra.close()
        item_cnt = len(rows)
        self.tw_reserve.setRowCount(item_cnt)
        style = "::section {""background-color: lightblue; }"
        self.tw_reserve.horizontalHeader().setStyleSheet(style)
        self.tw_reserve.resizeRowsToContents()

        #self.tw_reserve.setHorizontalScrollBar(scroll_bar)
        # self.tw_reserve.setShowGrid(False)

        #호실 정보 보여주기
        for idx, col in enumerate(rows):
            # self.tw_reserve.item(0, 0).setBackground(QBrush(Qt.red))

            self.tw_reserve.setItem(idx, 0, QTableWidgetItem(col["ILJA"]))#일자
            self.tw_reserve.setItem(idx, 1, QTableWidgetItem(col["WK"]))#요일
            self.tw_reserve.item(idx, 1).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            if col["WK"] == '일':
                self.tw_reserve.item(idx, 0).setForeground(QBrush(Qt.red))
                self.tw_reserve.item(idx, 1).setForeground(QBrush(Qt.red))
                self.tw_reserve.item(idx, 0).setFont(QFont("Arial", 12, QFont.Bold, italic=False))  # 글자 폰트 설정.
                self.tw_reserve.item(idx, 1).setFont(QFont("Arial", 12, QFont.Bold, italic=False))  # 글자 폰트 설정.
            elif col["WK"] == '토':
                self.tw_reserve.item(idx, 0).setForeground(QBrush(Qt.blue))
                self.tw_reserve.item(idx, 1).setForeground(QBrush(Qt.blue))
                self.tw_reserve.item(idx, 0).setFont(QFont("Arial", 12, QFont.Bold, italic=False))  # 글자 폰트 설정.
                self.tw_reserve.item(idx, 1).setFont(QFont("Arial", 12, QFont.Bold, italic=False))  # 글자 폰트 설정.
            else:
                self.tw_reserve.item(idx, 0).setForeground(QBrush(Qt.black))
                self.tw_reserve.item(idx, 1).setForeground(QBrush(Qt.black))
            self.tw_reserve.setItem(idx, 2, QTableWidgetItem(col["RN"]))#호실
            self.tw_reserve.item(idx, 2).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.tw_reserve.setItem(idx, 3, QTableWidgetItem(col["ROOM_TYPE"]))#객실타입
            self.tw_reserve.item(idx, 3).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.tw_reserve.setItem(idx, 4, QTableWidgetItem(col["RESERVE_NAME"]))#예약자
            self.tw_reserve.setItem(idx, 5, QTableWidgetItem(col["SITE_NAME"]))#예약사이트
            self.tw_reserve.setItem(idx, 6, QTableWidgetItem(col["RESERVE_ILJA"]))#예약일자
            self.tw_reserve.setItem(idx, 7, QTableWidgetItem(col["RESERVE_NUMBER"]))#예약번호

            self.tw_reserve.setItem(idx, 8, QTableWidgetItem(col["CANCEL_START_ILJA"])) # 취소사항 시작일자
            self.tw_reserve.item(idx, 8).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.tw_reserve.setItem(idx, 9, QTableWidgetItem(col["CANCEL_END_ILJA"])) # 취소사항 종료일자
            self.tw_reserve.item(idx, 9).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.tw_reserve.setItem(idx, 10, QTableWidgetItem(col["CANCEL_PAYMENT_NAME"])) # 취소사항 환불수단

            if col["CANCEL_SUM"] is None or col["CANCEL_SUM"] == '0' or col["CANCEL_SUM"] == 0:#환불금액
                self.tw_reserve.setItem(idx, 11, QTableWidgetItem('')) #환불금액
            else:
                self.tw_reserve.setItem(idx, 11, QTableWidgetItem(str(col["CANCEL_SUM"])+'원'))#환불금액

            # self.tw_reserve.setItem(idx, 11, QTableWidgetItem(str(col["CANCEL_SUM"])+'원'))#환불금액
            self.tw_reserve.item(idx, 11).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tw_reserve.setItem(idx, 12, QTableWidgetItem(col["CANCEL_BIGO"])) #취소 참고사항
            self.tw_reserve.setItem(idx, 13, QTableWidgetItem(col["MNT_ID"])) # 취소자 ID
            self.tw_reserve.item(idx, 13).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)

            self.tw_reserve.item(idx, 0).setBackground(QBrush(QColor(255, 255, 255)))
            self.tw_reserve.item(idx, 1).setBackground(QBrush(QColor(255, 255, 255)))
            self.tw_reserve.item(idx, 2).setBackground(QBrush(QColor(255, 255, 255)))
            self.tw_reserve.item(idx, 3).setBackground(QBrush(QColor(255, 255, 255)))
            self.tw_reserve.item(idx, 4).setBackground(QBrush(QColor(255, 255, 255)))
            self.tw_reserve.item(idx, 5).setBackground(QBrush(QColor(255, 255, 255)))
            self.tw_reserve.item(idx, 6).setBackground(QBrush(QColor(255, 255, 255)))
            self.tw_reserve.item(idx, 7).setBackground(QBrush(QColor(255, 255, 255)))
            self.tw_reserve.item(idx, 8).setBackground(QBrush(QColor(179, 255, 175)))
            self.tw_reserve.item(idx, 9).setBackground(QBrush(QColor(179, 255, 175)))
            self.tw_reserve.item(idx, 10).setBackground(QBrush(QColor(179, 255, 175)))
            self.tw_reserve.item(idx, 11).setBackground(QBrush(QColor(179, 255, 175)))
            self.tw_reserve.item(idx, 12).setBackground(QBrush(QColor(179, 255, 175)))
            self.tw_reserve.item(idx, 13).setBackground(QBrush(QColor(255, 255, 255)))
        self.tw_reserve.resizeColumnsToContents()#col 자릿수에 맡게 보여주려면...
        self.tw_reserve.verticalHeader().setVisible(False)  # row header 숨기기
        self.tw_reserve.setEditTriggers(QAbstractItemView.NoEditTriggers)#수정 불가하게
        self.tw_reserve.setSelectionBehavior(QAbstractItemView.SelectRows)#한줄씩 선택
        # self.tw_reserve.setCurrentCell(row_focus_cnt-1, 0)  # 한 ROW 위로 이동

    #죄회된 자료를 엑셀 파일로 만들려면.......
    def excel_rtn(self):
        dir = str(os.path.join(Path.home(), "Downloads")) + "\\"
        # dir = path_to_download_folder+"\\"
        filename = '예약취소현황'
        file_ext = '.xlsx'

        output_file_name = '%s%s%s' % (dir, filename, file_ext)
        uniq = 1
        while os.path.exists(output_file_name):
            output_file_name = '%s%s(%d)%s' % (dir, filename, uniq, file_ext)
            uniq += 1
        # print(output_file_name)


        #workbook 생성하기(1개의 시트가 생성된 상태)
        wb = Workbook()
        # 현재 workbook의 활성화 된 Sheet 가져오기
        ws = wb.active
        #테두리 선을 표현하려면
        box = Border(top=Side(border_style='thin', color='000000'),
                     right=Side(border_style='thin', color='000000'),
                     bottom=Side(border_style='thin', color='000000'),
                     left=Side(border_style='thin', color='000000'))

        if self.cb_site_sch.currentText() == '전체':
            site_code = '%'
        else:
            #코드 값 가져오기
            mra = mariadb_conn().conn
            csr = mra.cursor()
            query_sel = "SELECT FN_APPCOMP_CODE('"+site_name+"') "
            # print(query_sel)
            csr.execute(query_sel)
            rows = csr.fetchall()
            csr.close()
            mra.close()
            site_code = rows[0][0] #site code값을 가져온다

        start_ilja_sch = self.de_start_ilja_sch.date().toString('yyyy-MM-dd')#입실일자
        end_ilja_sch = self.de_end_ilja_sch.date().toString('yyyy-MM-dd')#퇴실일자

        #일자별 호실 정보 가져오기 60일 치
        mra = mariadb_conn().conn
        csr = mra.cursor(dictionary=True) #딕셔너리 형태로 표현하기 위해....
        query_sel = ("SELECT DATE_FORMAT(a.ROOM_ILJA,'%Y-%m-%d') as ILJA,(select WK from yeogak_reserve_day_tbl where ilja = a.room_ilja) AS WK,A.room_name AS RN, "
                    "(SELECT room_type from yeogak_room_name_info_tbl where comp_code = a.comp_code and room_name = a.room_name) "
                    "AS ROOM_TYPE, A.RESERVE_NAME,decode_oracle(A.RESERVE_SITE,'','',FN_APPCOMP_NAME(A.RESERVE_SITE)) AS SITE_NAME, "
                    "A.RESERVE_ILJA,A.RESERVE_NUMBER,A.RESERVE_START_ILJA,A.RESERVE_END_ILJA, decode_oracle(a.payment,'','',null,'',FN_CODE_NAME('PAY',a.payment)) AS PAYMENT_NAME, FORMAT(a.PAY_SUM,0) AS PAY_SUM,FORMAT(a.EXPEC_SUM,0) AS EXPEC_SUM, "
                    "DECODE_ORACLE(A.DORMITORY_SEX,'g','여자','m','남자','') AS DORMITORY_SEX,DECODE_ORACLE(A.DORMITORY_CLOSE,'c','Yes','') AS DORMITORY_CLOSE, "
                    "A.WORKER,A.BIGO, " 
                    "(SELECT decode_oracle(add_payment,'', '',FN_CODE_NAME('PAY',add_payment)) from yeogak_addpayment2025 WHERE comp_code = b.comp_code "
                    "AND key_ilja = a.key_ilja AND room_ilja =a.room_ilja) AS ADD_PAYMENT, "
                    "(SELECT decode_oracle(add_sum, NULL, '',FORMAT(add_sum,0)) "
                    "from yeogak_addpayment2025_tbl WHERE comp_code = a.comp_code AND key_ilja = a.key_ilja AND room_ilja =a.room_ilja) AS ADD_SUM, "
                    "(SELECT decode_oracle(add_bigo, NULL,'',add_bigo) from yeogak_addpayment2025_tbl WHERE comp_code = a.comp_code "
                    "AND key_ilja = a.key_ilja AND room_ilja =a.room_ilja) AS ADD_BIGO, "
                    "datediff(date_format(a.reserve_end_ilja, '%Y-%m-%d'),date_format(a.reserve_start_ilja, '%Y-%m-%d')) AS DATE_CHA, "                      
                    "B.CANCEL_START_ILJA, "
                    "B.CANCEL_END_ILJA,DECODE_ORACLE(B.CANCEL_PAYMENT,'','',FN_CODE_NAME('CPY',B.CANCEL_PAYMENT))  AS CANCEL_PAYMENT_NAME, "
                    "FORMAT(B.CANCEL_SUM,0) AS CANCEL_SUM, "
                    "B.CANCEL_BIGO,B.MNT_ID,A.SPECIAL_NOTE "
                    "FROM  yeogak_reserve2025_tbl a, yeogak_reserve2025_cancel_tbl b "   
                    "where a.comp_code = %s and a.room_ilja BETWEEN %s AND %s  and a.reserve_site LIKE %s AND a.cancel_chk = '1' " 							
                    "and a.comp_code = b.comp_code and a.key_ilja = b.key_ilja AND a.reserve_number = b.reserve_number "
                    "ORDER BY a.room_ilja,a.reserve_name ")
        # print(query_sel)
        t = (self.comp_code,start_ilja_sch,end_ilja_sch,site_code)
        csr.execute(query_sel,t)
        rows = csr.fetchall()
        csr.close()
        mra.close()
        item_cnt = len(rows)
        if item_cnt > 0: #조회된 결과가 없으면 메세지를 보낸다
            ws['A3'] = '일  자'  # 일자
            ws.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=1).border = box #테두리 선
            ws.cell(row=3, column=1).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['A'].width = 13 ## 칼럼 폭(열 가로 길이) 변경
            ws['B3'] = '요일'  # 요일
            ws.cell(row=3, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=2).border = box #테두리 선
            ws.cell(row=3, column=2).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws['C3'] = '호 실'  # 호실
            ws.cell(row=3, column=3).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=3).border = box #테두리 선
            ws.cell(row=3, column=3).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws['D3'] = '객실타입'  # 객실타입
            ws.cell(row=3, column=4).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=4).border = box #테두리 선
            ws.cell(row=3, column=4).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws['E3'] = '예약자'  # 예약자
            ws.cell(row=3, column=5).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=5).border = box #테두리 선
            ws.cell(row=3, column=5).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws['F3'] = '예약사이트'  # 예약사이트
            ws.cell(row=3, column=6).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=6).border = box #테두리 선
            ws.cell(row=3, column=6).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['F'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['G3'] = '예약일자'  # 예약일자
            ws.cell(row=3, column=7).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=7).border = box #테두리 선
            ws.cell(row=3, column=7).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['G'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['H3'] = '예약번호'  # 예약번호
            ws.cell(row=3, column=8).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=8).border = box #테두리 선
            ws.cell(row=3, column=8).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            #ws['I3'] = '예상금액'  # 예상금액
            #ws.cell(row=3, column=9).alignment = Alignment(horizontal="center", vertical="center")
            #ws.cell(row=3, column=9).border = box #테두리 선
            #ws.cell(row=3, column=9).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws['I3'] = '취소시작일자'  # 취소시작일자
            ws.cell(row=3, column=9).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=9).border = box #테두리 선
            ws.cell(row=3, column=9).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws['J3'] = '취소종료일자'  # 취소종료일자
            ws.cell(row=3, column=10).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=10).border = box #테두리 선
            ws.cell(row=3, column=10).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['J'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['K3'] = '환불수단'  # 환불수단
            ws.cell(row=3, column=11).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=11).border = box #테두리 선
            ws.cell(row=3, column=11).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws['L3'] = '환불금금액'  # 결재금액
            ws.cell(row=3, column=12).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=12).border = box #테두리 선
            ws.cell(row=3, column=12).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['L'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['M3'] = '취소사유'  # 취소사유
            ws.cell(row=3, column=13).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=13).border = box #테두리 선
            ws.cell(row=3, column=13).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['M'].width = 20 ## 칼럼 폭(열 가로 길이) 변경
            ws['N3'] = '취소자ID'  # 취소자ID
            ws.cell(row=3, column=14).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=14).border = box #테두리 선
            ws.cell(row=3, column=14).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['N'].width = 12 ## 칼럼 폭(열 가로 길이) 변경
            ws['O3'] = '예상금액'  # 예상금액
            ws.cell(row=3, column=15).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=15).border = box #테두리 선
            ws.cell(row=3, column=15).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['O'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['P3'] = '입실일자'  # 입실일자
            ws.cell(row=3, column=16).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=16).border = box #테두리 선
            ws.cell(row=3, column=16).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['P'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['Q3'] = '퇴실일자'  # 퇴실일자
            ws.cell(row=3, column=17).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=17).border = box #테두리 선
            ws.cell(row=3, column=17).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['Q'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['R3'] = '결재수단'  # 결재수단
            ws.cell(row=3, column=18).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=18).border = box #테두리 선
            ws.cell(row=3, column=18).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws['S3'] = '결재금액'  # 결재금액
            ws.cell(row=3, column=19).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=19).border = box #테두리 선
            ws.cell(row=3, column=19).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['S'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['T3'] = '근무자'  # 근무자
            ws.cell(row=3, column=20).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=20).border = box #테두리 선
            ws.cell(row=3, column=20).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['T'].width = 10 ## 칼럼 폭(열 가로 길이) 변경
            ws['U3'] = '참고사항'  # 결재금액
            ws.cell(row=3, column=21).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=21).border = box #테두리 선
            ws.cell(row=3, column=21).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['U'].width = 20 ## 칼럼 폭(열 가로 길이) 변경
            ws['V3'] = '추가결재수단'  # 추가결재수단
            ws.cell(row=3, column=22).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=22).border = box #테두리 선
            ws.cell(row=3, column=22).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['V'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['W3'] = '추가결재금액'  # 추가결재금액
            ws.cell(row=3, column=23).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=23).border = box #테두리 선
            ws.cell(row=3, column=23).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['W'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['X3'] = '추가 참고사항'  # 추가 참고사항
            ws.cell(row=3, column=24).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=24).border = box #테두리 선
            ws.cell(row=3, column=24).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['X'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['Y3'] = '남/여'  # 현금
            ws.cell(row=3, column=25).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=25).border = box #테두리 선
            ws.cell(row=3, column=25).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws['Z3'] = '마감'  # 마감
            ws.cell(row=3, column=26).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=26).border = box #테두리 선
            ws.cell(row=3, column=26).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상

            #호실 정보 보여주기
            for idx, col in enumerate(rows):
                idx_r = idx+4
                ws.cell(row=idx_r, column=1).value = col["ILJA"]
                ws.cell(row=idx_r, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=1).border = box #테두리 선
                ws.cell(row=idx_r, column=2).value = col["WK"]
                ws.cell(row=idx_r, column=2).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=2).border = box #테두리 선
                ws.cell(row=idx_r, column=3).value = col["RN"]
                ws.cell(row=idx_r, column=3).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=3).border = box #테두리 선
                ws.cell(row=idx_r, column=4).value = col["ROOM_TYPE"]
                ws.cell(row=idx_r, column=4).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=4).border = box #테두리 선
                ws.cell(row=idx_r, column=5).value = col["RESERVE_NAME"]
                ws.cell(row=idx_r, column=5).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=5).border = box #테두리 선
                ws.cell(row=idx_r, column=6).value = col["SITE_NAME"]
                ws.cell(row=idx_r, column=6).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=6).border = box #테두리 선
                ws.cell(row=idx_r, column=7).value = col["RESERVE_ILJA"]
                ws.cell(row=idx_r, column=7).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=7).border = box #테두리 선
                ws.cell(row=idx_r, column=8).value = col["RESERVE_NUMBER"]
                ws.cell(row=idx_r, column=8).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=8).border = box #테두리 선 str(col["EXPEC_SUM"])+'원'
                ws.cell(row=idx_r, column=9).value = col["CANCEL_START_ILJA"]
                ws.cell(row=idx_r, column=9).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=9).border = box #테두리 선
                ws.cell(row=idx_r, column=10).value =col["CANCEL_END_ILJA"]
                ws.cell(row=idx_r, column=10).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=10).border = box #테두리 선
                ws.cell(row=idx_r, column=11).value = col["CANCEL_PAYMENT_NAME"]
                ws.cell(row=idx_r, column=11).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=11).border = box #테두리 선
                ws.cell(row=idx_r, column=12).value = str(col["CANCEL_SUM"])+'원'
                ws.cell(row=idx_r, column=12).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=12).border = box #테두리 선
                ws.cell(row=idx_r, column=13).value = col["CANCEL_BIGO"]
                ws.cell(row=idx_r, column=13).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=13).border = box #테두리 선
                ws.cell(row=idx_r, column=14).value =col["MNT_ID"]
                ws.cell(row=idx_r, column=14).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=14).border = box #테두리 선
                ws.cell(row=idx_r, column=15).value = str(col["EXPEC_SUM"])+'원'
                ws.cell(row=idx_r, column=15).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=15).border = box #테두리 선
                ws.cell(row=idx_r, column=16).value = col["RESERVE_START_ILJA"]
                ws.cell(row=idx_r, column=16).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=16).border = box #테두리 선
                ws.cell(row=idx_r, column=17).value = col["RESERVE_END_ILJA"]
                ws.cell(row=idx_r, column=17).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=17).border = box #테두리 선
                ws.cell(row=idx_r, column=18).value = col["PAYMENT_NAME"]
                ws.cell(row=idx_r, column=18).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=18).border = box #테두리 선
                ws.cell(row=idx_r, column=19).value = str(col["PAY_SUM"])+'원'
                ws.cell(row=idx_r, column=19).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=19).border = box #테두리 선
                ws.cell(row=idx_r, column=20).value = col["WORKER"]
                ws.cell(row=idx_r, column=20).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=20).border = box #테두리 선
                ws.cell(row=idx_r, column=21).value = col["BIGO"]
                ws.cell(row=idx_r, column=21).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=21).border = box #테두리 선
                # ws['P9'] = '추가결재수단'  # 추가결재수단
                if col["ADD_PAYMENT"] is None or col["ADD_PAYMENT"] == '':  # 추가 결재 수단이 없다면
                    ws.cell(row=idx_r, column=22).value = ''
                    ws.cell(row=idx_r, column=22).border = box #테두리 선
                    ws.cell(row=idx_r, column=23).value = '0원'
                    ws.cell(row=idx_r, column=23).border = box #테두리 선
                    ws.cell(row=idx_r, column=24).value = ''
                    ws.cell(row=idx_r, column=24).border = box #테두리 선
                else:
                    ws.cell(row=idx_r, column=22).value = col["ADD_PAYMENT"]
                    ws.cell(row=idx_r, column=22).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=idx_r, column=22).border = box #테두리 선
                    ws.cell(row=idx_r, column=23).value = str(col["ADD_SUM"])+'원'
                    ws.cell(row=idx_r, column=23).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=idx_r, column=23).border = box #테두리 선
                    ws.cell(row=idx_r, column=24).value = col["ADD_BIGO"]
                    ws.cell(row=idx_r, column=24).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=idx_r, column=24).border = box #테두리 선
                ws.cell(row=idx_r, column=25).value = col["DORMITORY_SEX"]
                ws.cell(row=idx_r, column=25).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=25).border = box #테두리 선
                ws.cell(row=idx_r, column=26).value = col["DORMITORY_CLOSE"]
                ws.cell(row=idx_r, column=26).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=26).border = box #테두리 선
            wb.save(output_file_name)
            QMessageBox.about(self, "정보", output_file_name+" 파일이 저장되었습니다.")
        else:
            QMessageBox.about(self, "정보", "조회된 자료가 없음. 확인 바랍니다.")