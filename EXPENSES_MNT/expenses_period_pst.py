from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.QtCore import *
from PyQt5.QtGui import QBrush, QRegExpValidator, QColor, QPalette, QFont
from PyQt5.QtWidgets import *

import pandas as pd
import subprocess
import webbrowser

from config.mariadb_connection import mariadb_conn
import mariadb
from datetime import timedelta,datetime

from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,PatternFill,Border,Side,colors
from pathlib import Path
import os
import threading
import time

from function.inlink import inlink_url

# class StockIndexWindow(QMainWindow, form_class):
class Expenses_period_pst_Window(QMainWindow):
    def __init__(self,pf_os,comp_code,user_id):
        super().__init__()
        self.move(10, 50)
        self.setFixedSize(1736, 903)
        if pf_os == 'Windows':
            option_ui = 'UiDir/expenses_period_pst.ui'
        else:
            option_ui = 'UiDir/expenses_period_pst.ui'

        self.comp_code = comp_code
        self.user_id = user_id
        # option_ui = 'UiDir/StockCompInfo.ui'
        uic.loadUi(option_ui, self)

        regExp = QRegExp("[0-9]*") #edit에 숫자만 입력 처리하기 위해
        palette = QPalette()
        palette.setColor(QPalette.Highlight, QColor(144, 153, 234))  # default ==> Qt.darkBlue
        # palette.setColor(QPalette.HighlightedText, Qt.red)  # default ==> Qt.white
        self.tw_expenses.setPalette(palette)
        self.init_rtn()#오늘 일자를 먼저 보여주려면......
        self.pb_search.clicked.connect(self.select_rtn)  # 자료 조회
        self.pb_clear.clicked.connect(self.expenes_clear)  # 저장 화면을 클리어
        self.pb_excel.clicked.connect(self.excel_rtn)  #엑셀 파일로 저장 어
        self.tw_expenses.clicked.connect(self.mouse_click_rtn) # 클릭시 해당일자 호실의 정보를 상당 입력란에 보여주려면.....
        self.tw_expenses.keyPressEvent = self.keyPressEvent #UP, DOWN 키를 누를 때 해당일자 호실의 정보를 상당 입력란에 보여주려면.....
        self.tb_internet.clicked.connect(self.internet_sch)#링크 주소 인터넷으로 연결하기
        self.thread_rtn()#쓰레드 30초마다 자료를 다시 가져오려면

    def thread_rtn(self):
        self.first_sch()
        tt= threading.Timer(30, self.thread_rtn)

    #오늘 일자를 먼저 보여주려면....
    def init_rtn(self):
        # 현재 일자를 검색
        mra = mariadb_conn().conn
        csr = mra.cursor()
        query_sel = """SELECT date_format(date_add(CURDATE(), INTERVAL -31 day),'%Y-%m-%d'),
                       date_format(date_add(CURDATE(), INTERVAL 2 day),'%Y-%m-%d'),
                       date_format(CURDATE(),'%Y-%m-%d')  from dual """
        # 결제(query_sel)
        csr.execute(query_sel)
        rows = csr.fetchall()
        csr.close()
        mra.close()
        self.st_ilja_upno = rows[0][0]#현재일 보다 31일 전의 날짜이면 update를 못하게 한다면.....
        en_ilja = rows[0][1]
        self.now_date = rows[0][2]  # 현재일자를 가져고 검색 및 저장, 초기화에 활용한다.....

        self.de_start_ilja_sch.setDate(datetime.strptime(self.st_ilja_upno, '%Y-%m-%d'))  # 시작일자
        self.de_end_ilja_sch.setDate(datetime.strptime(en_ilja, '%Y-%m-%d'))  # 종료일자

    #인터넷 링크 주소 연결하기
    def internet_sch(self):
        inlink_url(self.ed_inlink.text())#인트넷으로 연결하기

    #조회시 일자 값이 정확한지 체크
    def select_rtn(self):
        if self.de_start_ilja_sch.date() > self.de_end_ilja_sch.date():
            QMessageBox.about(self, "정보", "일자를 정확히 입력하세요....")
            self.tw_expenses.setRowCount(0)
            self.de_start_ilja_sch.setFocus()
        else:
            self.first_sch()

    def expenes_clear(self):
        self.ed_ilja.setText("")#일자
        self.ed_name.setText("")  # 품명
        self.ed_qty.setText("")  # 수량
        self.ed_price.setText("")  # 단가
        self.ed_sum.setText("")  # 금액
        self.ed_po.setText("")  # 결재 금액
        self.ed_tel.setText("")  # 당일 근무자
        self.ed_bigo.setText("")  # 예약자
        self.ed_inlink.setText("")  # 예약자

    #KEY_UP, KEY_DOWN을 눌렀다면 .....
    def keyPressEvent (self, eventQKeyEvent):
        key = eventQKeyEvent.key()
        if key == Qt.Key_Up: #up 키를 누를 때마다  해당일자 호실의 정보를 상당 입력란에 보여주려면.....
            row = self.tw_expenses.currentRow() #현재 ROW 값을 가져오려면..
            if row == 0: # 처음 일때 값을 1로 하기 위해
                row += 1
            self.tw_expenses.setCurrentCell(row-1,0)  #한 ROW 위로 이동
            self.search_rtn()

        elif key == Qt.Key_Down:  # down 키를 누를 때마다  해당일자 호실의 정보를 상당 입력란에 보여주려면.....
            row = self.tw_expenses.currentRow() #현재 ROW 값을 가져오려면..
            if row >= (self.tw_expenses.rowCount() -1): #마지막 ROW일때 계속 마지막 해당일자 호실의 정보를 상당 입력란에 보여주려면.....
               row -= 1
            self.tw_expenses.setCurrentCell(row + 1, 0) #한 ROW 밑으로 이동
            self.search_rtn()

    # 클릭시 해당일자 호실의 정보를 상당 입력란에 보여주려면.....
    def mouse_click_rtn(self):
        row = self.tw_expenses.currentRow() #현재 ROW 값을 가져오려면..
        self.tw_expenses.setCurrentCell(row, 0)  # ROW 이동
        self.search_rtn()


    #해당일자 호실에 대한 정보를 상단 입력란에 뿌려주려면.....
    def search_rtn(self):
        self.ed_ilja.setText(self.tw_expenses.item(self.tw_expenses.currentRow(), 0).text())#일자
        self.ed_name.setText(self.tw_expenses.item(self.tw_expenses.currentRow(), 2).text())  # 품명
        self.ed_payment.setText(self.tw_expenses.item(self.tw_expenses.currentRow(), 3).text())  # 수량
        self.ed_qty.setText(self.tw_expenses.item(self.tw_expenses.currentRow(), 4).text())  # 수량
        self.ed_price.setText(self.tw_expenses.item(self.tw_expenses.currentRow(), 5).text())  # 단가
        self.ed_sum.setText(self.tw_expenses.item(self.tw_expenses.currentRow(), 6).text())  #금액
        self.ed_po.setText(self.tw_expenses.item(self.tw_expenses.currentRow(), 7).text()) #구입처
        self.ed_tel.setText(self.tw_expenses.item(self.tw_expenses.currentRow(), 8).text()) #전화번호
        self.ed_bigo.setText(self.tw_expenses.item(self.tw_expenses.currentRow(), 9).text()) #구입처
        self.ed_inlink.setText(self.tw_expenses.item(self.tw_expenses.currentRow(), 10).text()) #인트넷링크

    #자료를 가져와 화면에 뿌려준다.
    def first_sch(self):
        self.expenes_clear()

        start_ilja_sch = self.de_start_ilja_sch.date().toString('yyyy-MM-dd')#시작일자
        end_ilja_sch = self.de_end_ilja_sch.date().toString('yyyy-MM-dd')#종료일자

        if self.ed_name_search.text() is None or self.ed_name_search.text()=='':
            name_sch = '%'
        else:
            name_sch = '%'+self.ed_name_search.text()+'%'

        #기간 동안 총 지출 합계를 계산하려면.....
        mra = mariadb_conn().conn
        csr = mra.cursor()
        query_sel = ("SELECT DECODE_ORACLE(SUM(EXP_SUM), NULL,0,FORMAT(SUM(EXP_SUM),0)) AS TOT_SUM,"
                     "COUNT(*) FROM expenses_day_tbl "
                    "WHERE comp_code = %s AND exp_ilja BETWEEN %s AND %s and exp_name like %s ")
        # 결제(query_sel)
        t=(self.comp_code,start_ilja_sch,end_ilja_sch,name_sch)
        csr.execute(query_sel,t)
        rows = csr.fetchall()
        csr.close()
        mra.close()
        self.ed_ge_1.setText(start_ilja_sch+' ~ '+end_ilja_sch)#기간
        self.ed_ge_2.setText(str(rows[0][1])+'건')  # 건수
        self.ed_ge_3.setText(str(rows[0][0])+'원')  # 합계

        #기간 내에 지출 정보정보 가져오기
        mra = mariadb_conn().conn
        csr = mra.cursor(dictionary=True) #딕셔너리 형태로 표현하기 위해....
        query_sel = ("SELECT DATE_FORMAT(A.ILJA,'%Y-%m-%d') as ILJA,A.WK,B.EXP_ILJA,B.EXP_NAME,"
                     "decode_oracle(B.EXP_PAYMENT,'','',FN_CODE_NAME('PAY',B.EXP_PAYMENT)) AS EXP_PAYMENT,"
                    "FORMAT(B.EXP_QTY,0) AS EXP_QTY,FORMAT(B.EXP_PRICE,0) AS EXP_PRICE,FORMAT(B.EXP_SUM,0) AS EXP_SUM, "
                    "B.EXP_PO,EXP_TEL,B.EXP_BIGO,B.KEY_ILJA,B.EXP_INLINK "
                    "FROM (SELECT ilja,WK  FROM yeogak_reserve_day_tbl "
                    "WHERE ilja BETWEEN %s AND %s) a, "
                    "expenses_day_tbl b "
                    " where b.comp_code = %s AND a.ilja = b.exp_ilja and b.exp_name like %s "
                    "ORDER BY A.ILJA,B.KEY_ILJA ")
        # 결제(query_sel)
        t=(start_ilja_sch,end_ilja_sch,self.comp_code,name_sch)
        csr.execute(query_sel,t)
        rows = csr.fetchall()
        csr.close()
        mra.close()
        item_cnt = len(rows)
        self.tw_expenses.setRowCount(item_cnt)
        style = "::section {""background-color: lightblue; }"
        self.tw_expenses.horizontalHeader().setStyleSheet(style)
        # self.tw_reserve.setShowGrid(False)

        #호실 정보 보여주기
        for idx, col in enumerate(rows):
            # self.tw_reserve.item(0, 0).setBackground(QBrush(Qt.red))

            self.tw_expenses.setItem(idx, 0, QTableWidgetItem(col["ILJA"]))#일자
            self.tw_expenses.setItem(idx, 1, QTableWidgetItem(col["WK"]))#요일
            self.tw_expenses.item(idx, 1).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            if col["WK"] == '일':
                self.tw_expenses.item(idx, 0).setForeground(QBrush(Qt.red))
                self.tw_expenses.item(idx, 1).setForeground(QBrush(Qt.red))
                self.tw_expenses.item(idx, 0).setFont(QFont("Arial", 12, QFont.Bold, italic=False))  # 글자 폰트 설정.
                self.tw_expenses.item(idx, 1).setFont(QFont("Arial", 12, QFont.Bold, italic=False))  # 글자 폰트 설정.
            elif col["WK"] == '토':
                self.tw_expenses.item(idx, 0).setForeground(QBrush(Qt.blue))
                self.tw_expenses.item(idx, 1).setForeground(QBrush(Qt.blue))
                self.tw_expenses.item(idx, 0).setFont(QFont("Arial", 12, QFont.Bold, italic=False))  # 글자 폰트 설정.
                self.tw_expenses.item(idx, 1).setFont(QFont("Arial", 12, QFont.Bold, italic=False))  # 글자 폰트 설정.
            else:
                self.tw_expenses.item(idx, 0).setForeground(QBrush(Qt.black))
                self.tw_expenses.item(idx, 1).setForeground(QBrush(Qt.black))

            self.tw_expenses.setItem(idx, 2, QTableWidgetItem(col["EXP_NAME"]))#품명
            self.tw_expenses.setItem(idx, 3, QTableWidgetItem(col["EXP_PAYMENT"]))#결재수단
            if col["EXP_QTY"] is None or col["EXP_QTY"] == 0 or col["EXP_QTY"] == '0':
                self.tw_expenses.setItem(idx, 4, QTableWidgetItem(''))  # 수량
                self.tw_expenses.setItem(idx, 5, QTableWidgetItem(''))  # 단가
                self.tw_expenses.setItem(idx, 6, QTableWidgetItem(''))  # 금액
            else:
                self.tw_expenses.setItem(idx, 4, QTableWidgetItem(str(col["EXP_QTY"])))#수량
                self.tw_expenses.item(idx, 4).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tw_expenses.setItem(idx, 5, QTableWidgetItem(str(col["EXP_PRICE"])+'원'))#단가
                self.tw_expenses.item(idx, 5).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tw_expenses.setItem(idx, 6, QTableWidgetItem(str(col["EXP_SUM"])+'원'))#금액
                self.tw_expenses.item(idx, 6).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tw_expenses.setItem(idx, 7, QTableWidgetItem(col["EXP_PO"]))#구입처
            self.tw_expenses.setItem(idx, 8, QTableWidgetItem(col["EXP_TEL"]))#전화번호
            self.tw_expenses.setItem(idx, 9, QTableWidgetItem(col["EXP_BIGO"]))#비고
            self.tw_expenses.setItem(idx, 10, QTableWidgetItem(col["EXP_INLINK"]))#비고

        self.tw_expenses.resizeRowsToContents()
        self.tw_expenses.resizeColumnsToContents()#col 자릿수에 맡게 보여주려면...
        self.tw_expenses.verticalHeader().setVisible(False)  # row header 숨기기
        self.tw_expenses.setEditTriggers(QAbstractItemView.NoEditTriggers)#수정 불가하게
        self.tw_expenses.setSelectionBehavior(QAbstractItemView.SelectRows)#한줄씩 선택
        self.tw_expenses.setCurrentCell(item_cnt-1, 0)  # 마지막 ROW 위로 이동
        # if item_cnt > 5 :
        #     self.tw_expenses.setCurrentCell(item_cnt, 0)  # 한 ROW 위로 이동
            # self.tw_expenses.setCurrentCell(item_cnt-4, 0)  # 한 ROW 위로 이동

    #엑셀 파일을 만들려면....
    def excel_rtn(self):
        dir = str(os.path.join(Path.home(), "Downloads")) + "\\"
        # dir = path_to_download_folder+"\\"
        filename = '기간별지출현황'
        file_ext = '.xlsx'

        output_file_name = '%s%s%s' % (dir, filename, file_ext)
        uniq = 1
        while os.path.exists(output_file_name):
            output_file_name = '%s%s(%d)%s' % (dir, filename, uniq, file_ext)
            uniq += 1
        # 결제(output_file_name)


        #workbook 생성하기(1개의 시트가 생성된 상태)
        wb = Workbook()
        # 현재 workbook의 활성화 된 Sheet 가져오기
        ws = wb.active
        #테두리 선을 표현하려면
        box = Border(top=Side(border_style='thin', color='000000'),
                     right=Side(border_style='thin', color='000000'),
                     bottom=Side(border_style='thin', color='000000'),
                     left=Side(border_style='thin', color='000000'))


        start_ilja_sch = self.de_start_ilja_sch.date().toString('yyyy-MM-dd')#시작일자
        end_ilja_sch = self.de_end_ilja_sch.date().toString('yyyy-MM-dd')#종료일자

        if self.ed_name_search.text() is None or self.ed_name_search.text()=='':
            name_sch = '%'
            name_sch_title = ''
        else:
            name_sch = '%'+self.ed_name_search.text()+'%'
            name_sch_title = ',   품 명 : ' +self.ed_name_search.text()

        #기간 동안 총 지출 합계를 계산하려면.....
        mra = mariadb_conn().conn
        csr = mra.cursor()
        query_sel = ("SELECT DECODE_ORACLE(SUM(EXP_SUM), NULL,0,FORMAT(SUM(EXP_SUM),0)) AS TOT_SUM,"
                     "COUNT(*) FROM expenses_day_tbl "
                    "WHERE comp_code = %s AND exp_ilja BETWEEN %s AND %s and exp_name like %s ")
        # 결제(query_sel)
        t=(self.comp_code,start_ilja_sch,end_ilja_sch,name_sch)
        csr.execute(query_sel,t)
        rows = csr.fetchall()
        csr.close()
        mra.close()

        ws.title ='지출 현황('+str(start_ilja_sch)+'~'+str(end_ilja_sch)+')' # 해당 sheet의 sheet명 변경하기
        #조회 조건을 엑셀 파일 첫 줄에 표현하려면....
        ws.merge_cells('A1:J1')  # 조회 조건
        # ws['A1'] = '조회 조건 => 일자 : '+str(start_ilja_sch)+'~'+str(end_ilja_sch)+',  호실 : '+self.cb_room_name_sch.currentText()+',  사이트 : '+self.cb_site_sch.currentText()
        ws.cell(row=1, column=1).value = '조회 조건 => 일자 : '+str(start_ilja_sch)+'~'+str(end_ilja_sch)+name_sch_title
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells('A3:C3') #지출 현황
        ws.cell(row=3, column=1).value ='지   출   현   황'
        ws.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=3, column=1).border = box #테두리 선
        ws.cell(row=3, column=1).fill = PatternFill(patternType='solid',fgColor='daf7da') #색상

        ws['A4'] = '기   간' # 기간
        ws.cell(row=4, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=4, column=1).border = box #테두리 선
        ws.cell(row=4, column=1).fill = PatternFill(patternType='solid',fgColor='daf7da') #색상

        ws['B4'] = '건 수'  # 건수
        ws.cell(row=4, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=4, column=2).border = box #테두리 선
        ws.cell(row=4, column=2).fill = PatternFill(patternType='solid',fgColor='daf7da') #색상

        ws['C4'] = '지출 금액'  # 지출금액
        ws.cell(row=4, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=4, column=3).border = box #테두리 선
        ws.cell(row=4, column=3).fill = PatternFill(patternType='solid',fgColor='daf7da') #색상

        ws['A5'] = start_ilja_sch+' ~ '+end_ilja_sch #기간
        ws.cell(row=5, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=5, column=1).border = box
        ws.cell(row=5, column=1).font=Font(color='0000ff',size=11)
        ws['B5'] = str(rows[0][1])+'건' #건수
        ws.cell(row=5, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=5, column=2).border = box
        ws.cell(row=5, column=2).font=Font(color='0000ff',size=11)
        ws['C5'] = str(rows[0][0])+'원' #지출금액
        ws.cell(row=5, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=5, column=3).border = box
        ws.cell(row=5, column=3).font=Font(color='0000ff',size=11)

        #기간 내에 지출 정보정보 가져오기
        mra = mariadb_conn().conn
        csr = mra.cursor(dictionary=True) #딕셔너리 형태로 표현하기 위해....
        query_sel = ("SELECT DATE_FORMAT(A.ILJA,'%Y-%m-%d') as ILJA,A.WK,B.EXP_ILJA,B.EXP_NAME,"
                     "decode_oracle(B.EXP_PAYMENT,'','',FN_CODE_NAME('PAY',B.EXP_PAYMENT)) AS EXP_PAYMENT,"
                    "FORMAT(B.EXP_QTY,0) AS EXP_QTY,FORMAT(B.EXP_PRICE,0) AS EXP_PRICE,FORMAT(B.EXP_SUM,0) AS EXP_SUM, "
                    "B.EXP_PO,EXP_TEL,B.EXP_BIGO,B.KEY_ILJA,B.EXP_INLINK "
                    "FROM (SELECT ilja,WK  FROM yeogak_reserve_day_tbl "
                    "WHERE ilja BETWEEN %s AND %s) a, "
                    "expenses_day_tbl b "
                    " where b.comp_code = %s AND a.ilja = b.exp_ilja and b.exp_name like %s "
                    "ORDER BY A.ILJA,B.KEY_ILJA ")
        # 결제(query_sel)
        t=(start_ilja_sch,end_ilja_sch,self.comp_code,name_sch)
        csr.execute(query_sel,t)
        rows = csr.fetchall()
        csr.close()
        mra.close()
        item_cnt = len(rows)
        if item_cnt > 0: #조회된 결과가 없으면 메세지를 보낸다
            ws['A8'] = '지출 일자'  # 일자
            ws.cell(row=8, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=8, column=1).border = box #테두리 선
            ws.cell(row=8, column=1).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['A'].width = 25 ## 칼럼 폭(열 가로 길이) 변경
            ws['B8'] = '요일'  # 요일
            ws.cell(row=8, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=8, column=2).border = box #테두리 선
            ws.cell(row=8, column=2).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws['C8'] = '품         명'  # 품명
            ws.cell(row=8, column=3).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=8, column=3).border = box #테두리 선
            ws.cell(row=8, column=3).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['C'].width = 30 ## 칼럼 폭(열 가로 길이) 변경
            ws['D8'] = '결재수단'  # 결재수단
            ws.cell(row=8, column=4).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=8, column=4).border = box #테두리 선
            ws.cell(row=8, column=4).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['D'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['E8'] = '수 량'  # 수량
            ws.cell(row=8, column=5).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=8, column=5).border = box #테두리 선
            ws.cell(row=8, column=5).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws['F8'] = '단    가'  # 단가
            ws.cell(row=8, column=6).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=8, column=6).border = box #테두리 선
            ws.cell(row=8, column=6).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['F'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['G8'] = '금    액'  # 금액
            ws.cell(row=8, column=7).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=8, column=7).border = box #테두리 선
            ws.cell(row=8, column=7).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['G'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['H8'] = '구 입 처'  # 구입처
            ws.cell(row=8, column=8).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=8, column=8).border = box #테두리 선
            ws.cell(row=8, column=8).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['H'].width = 25 ## 칼럼 폭(열 가로 길이) 변경
            ws['I8'] = '전화번호'  # 전화번호
            ws.cell(row=8, column=9).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=8, column=9).border = box #테두리 선
            ws.cell(row=8, column=9).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['I'].width = 15 ## 칼럼 폭(열 가로 길이) 변경
            ws['J8'] = '비         고'  # 예상금액
            ws.cell(row=8, column=10).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=8, column=10).border = box #테두리 선
            ws.cell(row=8, column=10).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['I'].width = 20 ## 칼럼 폭(열 가로 길이) 변경
            ws['K8'] = '링크주소'  # 링크주소
            ws.cell(row=8, column=11).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=8, column=11).border = box #테두리 선
            ws.cell(row=8, column=11).fill = PatternFill(patternType='solid',fgColor='c8c8cc') #색상
            ws.column_dimensions['K'].width = 70 ## 칼럼 폭(열 가로 길이) 변경

            #호실 정보 보여주기
            for idx, col in enumerate(rows):
                idx_r = idx+9
                # ws['A9'] = '일  자'  # 일자
                ws.cell(row=idx_r, column=1).value = col["ILJA"]
                ws.cell(row=idx_r, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=1).border = box #테두리 선
                ws.cell(row=idx_r, column=2).value = col["WK"]
                ws.cell(row=idx_r, column=2).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=2).border = box #테두리 선
                ws.cell(row=idx_r, column=3).value = col["EXP_NAME"]
                ws.cell(row=idx_r, column=3).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=idx_r, column=3).border = box #테두리 선
                ws.cell(row=idx_r, column=4).value = col["EXP_PAYMENT"]
                ws.cell(row=idx_r, column=4).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=idx_r, column=4).border = box #테두리 선
                ws.cell(row=idx_r, column=5).value = str(col["EXP_QTY"])
                ws.cell(row=idx_r, column=5).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=idx_r, column=5).border = box #테두리 선
                ws.cell(row=idx_r, column=6).value = str(col["EXP_PRICE"])+'원'
                ws.cell(row=idx_r, column=6).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=idx_r, column=6).border = box #테두리 선
                ws.cell(row=idx_r, column=7).value = str(col["EXP_SUM"])+'원'
                ws.cell(row=idx_r, column=7).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=idx_r, column=7).border = box #테두리 선
                ws.cell(row=idx_r, column=8).value = col["EXP_PO"]
                ws.cell(row=idx_r, column=8).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=idx_r, column=8).border = box #테두리 선
                ws.cell(row=idx_r, column=9).value = col["EXP_TEL"]
                ws.cell(row=idx_r, column=9).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=idx_r, column=9).border = box #테두리 선
                ws.cell(row=idx_r, column=10).value = col["EXP_BIGO"]
                ws.cell(row=idx_r, column=10).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=idx_r, column=10).border = box #테두리 선
                ws.cell(row=idx_r, column=11).value = col["EXP_INLINK"]
                ws.cell(row=idx_r, column=11).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=idx_r, column=11).border = box #테두리 선
            wb.save(output_file_name)
            QMessageBox.about(self, "정보", output_file_name+" 파일이 저장되었습니다.")
        else:
            QMessageBox.about(self, "정보", "조회된 자료가 없음. 확인 바랍니다.")